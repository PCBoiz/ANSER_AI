"""
src/core/inventory_import.py — đọc bảng TỔNG HỢP TỒN KHO thành InventoryLine.

VÌ SAO KHÔNG PHẢI OCR / VLM
---------------------------
Bản khách gửi là PDF, nhưng PDF đó được IN RA TỪ Excel — MISA/Fast/Bravo đều
xuất Excel gốc. Xin file .xlsx luôn thì đọc được chính xác 100%, còn OCR bảng
số tiền là tự chuốc lấy rủi ro đọc nhầm 5 thành 6 ở cột giá vốn. VLM để dành
cho thứ chỉ tồn tại dưới dạng ảnh: hoá đơn giấy của nhà xe.

Module này nhận `list[list]` nên nguồn nào cũng nạp được: openpyxl, csv,
pandas.values.tolist(), hay pdfplumber.extract_table() nếu buộc phải đọc PDF.

BA LỚP TỰ KIỂM (đây mới là phần quan trọng)
-------------------------------------------
Lỗi tệ nhất của mọi trình đọc bảng là LỆCH CỘT: giá trị nhảy sang ô bên cạnh,
mọi con số vẫn "hợp lý" nhưng sai hết. Ba lớp chặn:

1. Đơn giá BQ có sẵn trong file phải khớp với giá_trị / số_lượng do ta tính.
2. Dòng "Tổng cộng" của chính file phải khớp với tổng ta cộng lại.
3. ĐK + Nhập − Xuất = CK trên từng dòng (audit_inventory kiểm tiếp).

Không lớp nào tự sửa dữ liệu. Lệch thì báo, để người quyết định.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from typing import Any, Optional

from src.core.inventory import InventoryLine

# Dung sai khi đối chiếu — kế toán làm tròn đến đồng.
_VALUE_TOL = 2.0
_UNIT_TOL = 1.0
_QTY_TOL = 0.01

_DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")
# '25.000', '1.500.000' — mọi nhóm sau dấu chấm đều đúng 3 chữ số.
_THOUSAND_GROUPS_RE = re.compile(r"[+-]?\d{1,3}(?:\.\d{3})+")
_WAREHOUSE_RE = re.compile(r"kho\s*:\s*([^,\n]+)", re.IGNORECASE)

# (nhóm cột, cột con) -> tên trường trong InventoryLine
_GROUPS = {"dau ky": "opening", "nhap": "in", "xuat": "out", "cuoi ky": "closing"}
_SUBS = {"so luong": "qty", "gia tri": "value", "don gia": "unit"}


def _norm(value: Any) -> str:
    """Bỏ dấu, gộp khoảng trắng, hạ chữ thường — để so khớp tên cột."""
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"\s+", " ", text).strip().lower()


def parse_vn_number(value: Any) -> Optional[float]:
    """
    Số theo quy ước Việt Nam: '.' phân nhóm nghìn, ',' phân thập phân.

        '57.660,45' -> 57660.45      '(21,00)' -> -21.0
        '25.000'    -> 25000.0       '25.5'    -> 25.5
        '0,00'      -> 0.0           ''        -> None

    Âm ghi bằng ngoặc đơn (chuẩn kế toán). Bản xuất PDF hay rơi vãi dấu ')'
    thừa giữa các ô — ngoặc chỉ được coi là dấu âm khi bọc CẢ HAI đầu.

    DẤU CHẤM ĐƠN ĐỘC — chỗ dễ sai chết người. '25.000' người Việt viết là hai
    mươi lăm nghìn, nhưng `float()` đọc thành 25.0. Với ô giá dầu thì sai lệch
    1000 lần đó đi thẳng vào `fuel_ratio` và làm hỏng toàn bộ hiệu chỉnh mà
    không báo một chữ nào. Nên: dấu chấm chỉ được coi là phân nhóm nghìn khi
    các nhóm sau nó ĐỀU đúng 3 chữ số ('25.000', '1.500.000'); ngược lại nó là
    dấu thập phân ('25.5', '0.75').

    Ô rỗng trả None ("chưa biết"), KHÔNG trả 0.0 — nhầm hai thứ này là cách
    nhanh nhất để bịa ra một con số tài chính.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(" ", " ").replace(" ", "")
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    if text.startswith("-"):
        negative, text = True, text[1:]

    text = text.strip("()")
    if not text or not any(c.isdigit() for c in text):
        return None

    if "," in text:
        # Có dấu phẩy -> phẩy là thập phân, chấm là phân nhóm nghìn.
        text = text.replace(".", "").replace(",", ".")
    elif _THOUSAND_GROUPS_RE.fullmatch(text):
        text = text.replace(".", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


@dataclass
class ParseResult:
    lines: list[InventoryLine] = field(default_factory=list)
    warehouse: str = ""
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    warnings: list[str] = field(default_factory=list)
    checks: dict[str, Any] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        """
        Đọc sạch: có dòng, không lớp tự kiểm nào báo lệch, VÀ không thiếu cột
        thiết yếu.

        Vế cuối học được từ lần chạy thật đầu tiên (30/07/2026): parser đọc đúng
        tên kho, kỳ, mã hàng, tên hàng nhưng KHÔNG map được cột số nào, rồi vẫn
        trả ok=True vì `mismatches` rỗng — mà `mismatches` rỗng chính là *vì*
        không có cột nào để đối chiếu. Im lặng đúng lúc cần hét to nhất.
        """
        return (
            bool(self.lines)
            and not self.checks.get("mismatches")
            and not self.checks.get("missing_columns")
        )


def _is_code_label(text: str) -> bool:
    return text.startswith("ma hang") or text in {"ma vt", "ma sp", "ma hh"}


def _row_has(row: list[Any], names: tuple[str, ...]) -> bool:
    return any(any(_norm(c).startswith(n) for n in names) for c in row)


@dataclass
class _Header:
    label_i: int      # dòng chứa "Mã hàng" / "Tên hàng" / "ĐVT"
    group_i: int      # dòng chứa "Đầu kỳ" / "Nhập kho" / "Xuất kho" / "Cuối kỳ"
    sub_i: int        # dòng chứa "Số lượng" / "Giá trị" / "Đơn giá BQ"

    @property
    def data_start(self) -> int:
        return max(self.label_i, self.group_i, self.sub_i) + 1


def _find_header(rows: list[list[Any]]) -> Optional[_Header]:
    """
    Định vị header, chấp nhận cả hai cách bố trí đã gặp ngoài đời.

    MISA xuất kiểu A — nhóm nằm CÙNG dòng với "Mã hàng", cột con ở dòng KẾ TIẾP:

        [7] Tên kho | Mã hàng | Tên hàng | ĐVT | Đầu kỳ  |         |
        [8]         |         |          |     | Số lượng| Giá trị | Đơn giá BQ

    Kiểu B — nhóm ở dòng TRÊN, cột con cùng dòng với "Mã hàng":

        [5]         |         |          |     | Đầu kỳ  |         |
        [6] Tên kho | Mã hàng | Tên hàng | ĐVT | Số lượng| Giá trị | Đơn giá BQ

    Đoán sai một dòng là mất sạch cột số mà vẫn đọc được mã hàng — tức là ra
    một bảng trông bình thường với mọi con số bằng 0.
    """
    groups = tuple(_GROUPS)
    subs = tuple(_SUBS)

    for i, row in enumerate(rows):
        if not any(_is_code_label(_norm(c)) for c in row):
            continue
        below = rows[i + 1] if i + 1 < len(rows) else []
        above = rows[i - 1] if i > 0 else []

        sub_i = i if _row_has(row, subs) else (i + 1 if _row_has(below, subs) else i)
        group_i = i if _row_has(row, groups) else (i - 1 if _row_has(above, groups) else i)
        return _Header(label_i=i, group_i=group_i, sub_i=sub_i)
    return None


def _build_columns(rows: list[list[Any]], hdr: _Header) -> dict[str, int]:
    """
    Ánh xạ tên trường -> chỉ số cột.

    Ô gộp (merge) để lại ô trống ở các cột sau, nên tên nhóm phải được kéo sang
    phải (forward-fill) cho tới khi gặp nhóm mới.
    """
    label_row = rows[hdr.label_i]
    group_row = rows[hdr.group_i]
    sub_row = rows[hdr.sub_i]
    width = max(len(label_row), len(group_row), len(sub_row))

    cols: dict[str, int] = {}
    current = ""
    for idx in range(width):
        raw_group = _norm(group_row[idx]) if idx < len(group_row) else ""
        for key, short in _GROUPS.items():
            if raw_group.startswith(key):
                current = short
                break
        else:
            # Ô có chữ nhưng không phải tên nhóm. Nếu là tên cột con thì vẫn
            # thuộc nhóm đang mở (header một tầng); ngược lại đã ra khỏi vùng
            # nhóm nên đóng lại, đừng kéo nhầm sang cột kế bên.
            if raw_group and not any(raw_group.startswith(s) for s in _SUBS):
                current = ""

        # Nhãn cột lấy từ dòng nhãn; một số bản xuất để nhãn ở dòng cột con.
        for candidate in (label_row, sub_row):
            label = _norm(candidate[idx]) if idx < len(candidate) else ""
            if _is_code_label(label):
                cols.setdefault("code", idx)
            elif label.startswith("ten hang"):
                cols.setdefault("name", idx)
            elif label in {"dvt", "don vi tinh"}:
                cols.setdefault("unit", idx)
            elif label.startswith("ten kho"):
                cols.setdefault("warehouse", idx)

        sub = _norm(sub_row[idx]) if idx < len(sub_row) else ""
        if current:
            for key, short in _SUBS.items():
                if sub.startswith(key):
                    cols.setdefault(f"{current}_{short}", idx)
                    break
    return cols


def _cell(row: list[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_inventory_table(rows: list[list[Any]]) -> ParseResult:
    """
    Đọc bảng TỔNG HỢP TỒN KHO đã ở dạng ma trận ô.

    Không tự sửa gì. Mọi nghi ngờ đi vào `warnings` / `checks` để
    audit_inventory và người đọc quyết định.
    """
    res = ParseResult()
    rows = [r for r in rows if r is not None]

    # --- siêu dữ liệu: tên kho + khoảng thời gian --------------------------
    blob = "\n".join(" ".join("" if c is None else str(c) for c in r) for r in rows[:12])
    if (m := _WAREHOUSE_RE.search(blob)):
        res.warehouse = m.group(1).strip()
    dates = _DATE_RE.findall(blob)
    if len(dates) >= 2:
        (d0, m0, y0), (d1, m1, y1) = dates[0], dates[1]
        res.period_start = f"{y0}-{int(m0):02d}-{int(d0):02d}"
        res.period_end = f"{y1}-{int(m1):02d}-{int(d1):02d}"
    else:
        res.warnings.append(
            "Không đọc được khoảng thời gian của kỳ — các kiểm tra phụ thuộc số "
            "ngày (hàng chết, hàng bán chậm) sẽ bị bỏ qua."
        )

    hdr = _find_header(rows)
    if hdr is None:
        res.warnings.append("Không tìm thấy dòng tiêu đề có cột 'Mã hàng'. Bảng sai định dạng?")
        res.checks = {"missing_columns": ["header"], "mismatches": []}
        return res

    cols = _build_columns(rows, hdr)
    if "code" not in cols:
        res.warnings.append("Không xác định được cột 'Mã hàng'.")
        res.checks = {"missing_columns": ["code"], "mismatches": []}
        return res

    missing = [k for k in ("opening_qty", "in_qty", "out_qty", "closing_qty")
               if k not in cols]
    if missing:
        res.warnings.append(
            f"Thiếu cột số lượng: {', '.join(missing)}. Bản xuất có thể bị cắt cột "
            "(PDF nhiều trang thường tách Xuất kho / Cuối kỳ sang trang riêng)."
        )

    # --- dòng dữ liệu -------------------------------------------------------
    totals_row: Optional[list[Any]] = None
    unit_mismatches: list[dict[str, Any]] = []

    for row in rows[hdr.data_start:]:
        code = _cell(row, cols.get("code"))
        code_text = "" if code is None else str(code).strip()
        first = _norm(row[0] if row else "")

        if first.startswith("tong cong") or _norm(code_text).startswith("tong cong"):
            totals_row = row
            continue
        if not code_text:
            continue

        line = InventoryLine(
            code=code_text,
            name=str(_cell(row, cols.get("name")) or "").strip(),
            unit=str(_cell(row, cols.get("unit")) or "").strip(),
        )
        for stage in ("opening", "in", "out", "closing"):
            qty = parse_vn_number(_cell(row, cols.get(f"{stage}_qty")))
            val = parse_vn_number(_cell(row, cols.get(f"{stage}_value")))
            setattr(line, f"{stage}_qty", 0.0 if qty is None else qty)
            setattr(line, f"{stage}_value", val)

            # Lớp tự kiểm 1: đơn giá BQ trong file vs giá trị/số lượng ta tính.
            stated = parse_vn_number(_cell(row, cols.get(f"{stage}_unit")))
            if stated and val is not None and qty and abs(qty) > _QTY_TOL:
                computed = val / qty
                if abs(computed - stated) > max(_UNIT_TOL, abs(stated) * 0.001):
                    unit_mismatches.append({
                        "mã": line.code, "cột": stage,
                        "đơn_giá_trong_file": stated,
                        "đơn_giá_tính_lại": round(computed, 2),
                    })
        res.lines.append(line)

    if not res.lines:
        res.warnings.append("Không đọc được dòng hàng nào sau tiêu đề.")
        res.checks = {"missing_columns": [], "mismatches": []}
        return res

    res.checks = _verify(res, totals_row, cols, unit_mismatches)
    res.checks["missing_columns"] = missing
    if res.checks["mismatches"]:
        res.warnings.append(
            f"{len(res.checks['mismatches'])} điểm không khớp khi tự kiểm — nhiều "
            "khả năng bảng bị LỆCH CỘT lúc đọc. Đối chiếu lại trước khi dùng số."
        )
    return res


def _verify(
    res: ParseResult,
    totals_row: Optional[list[Any]],
    cols: dict[str, int],
    unit_mismatches: list[dict[str, Any]],
) -> dict[str, Any]:
    """Lớp tự kiểm 1 (đơn giá) + lớp 2 (dòng Tổng cộng của chính file)."""
    mismatches: list[dict[str, Any]] = []
    mismatches += [{"loại": "đơn_giá", **m} for m in unit_mismatches]

    totals_checked: dict[str, Any] = {}
    if totals_row is not None:
        for stage in ("opening", "in", "out", "closing"):
            for kind, tol in (("qty", _QTY_TOL), ("value", _VALUE_TOL)):
                idx = cols.get(f"{stage}_{kind}")
                stated = parse_vn_number(_cell(totals_row, idx))
                if stated is None:
                    continue
                vals = [getattr(ln, f"{stage}_{kind}") for ln in res.lines]
                computed = sum(v for v in vals if v is not None)
                totals_checked[f"{stage}_{kind}"] = {
                    "trong_file": stated, "tính_lại": round(computed, 2),
                }
                if abs(computed - stated) > max(tol, abs(stated) * 1e-6):
                    mismatches.append({
                        "loại": "tổng_cộng", "cột": f"{stage}_{kind}",
                        "trong_file": stated, "tính_lại": round(computed, 2),
                    })
    else:
        totals_checked["note"] = (
            "Bản xuất không có dòng 'Tổng cộng' — mất một lớp đối chiếu."
        )

    return {
        "rows_parsed": len(res.lines),
        "totals": totals_checked,
        "mismatches": mismatches,
    }


def load_xlsx(path: str, sheet: Optional[str] = None) -> ParseResult:
    """
    Đọc thẳng file Excel do MISA/Fast/Bravo xuất ra.

    openpyxl là phụ thuộc tuỳ chọn — chỉ cần khi dùng hàm này, nên import
    trong thân hàm để phần còn lại của module chạy được ở môi trường tối giản.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - phụ thuộc tuỳ chọn
        raise RuntimeError(
            "Cần openpyxl để đọc .xlsx:  pip install openpyxl"
        ) from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return parse_inventory_table(rows)


__all__ = ["ParseResult", "parse_inventory_table", "parse_vn_number", "load_xlsx"]
